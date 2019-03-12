import openpyxl as xl
import pandas as pd


def CBD():
    wb = xl.load_workbook('LIBOR.xlsx')
    hl = wb["Bank Holidays"]
    holidays = []
    row = 1
    col = 1
    while hl.cell(row, col).value is not None:
        holidays.append(hl.cell(row, col).value)
        row += 1
    return pd.offsets.CustomBusinessDay(holidays=holidays)
