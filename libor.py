import pandas as pd
import numpy as np
import openpyxl as xl
import statistics as stats
from datetime import date, datetime
from toolbox import CBD
import math
import urllib
import os


BASE_RATES = None


def create_file(url, filename):
    response = urllib.request.urlopen(url)
    html = response.read()

    with open(filename, 'wb') as f:
        f.write(html)


def update_file(url, filename):
    if(os.path.isfile(filename)):
        file_date = datetime.fromtimestamp(
            os.path.getctime(filename))
        if file_date.date() != date.today():
            create_file(url, filename)
    else:
        create_file(url, filename)


def calculate_libor_dist():

    # scrapes historic boe base rates
    br_url = 'https://www.bankofengland.co.uk/boeapps/iadb/fromshowcolumns' \
        '.asp?Travel=NIxRPxSUx&FromSeries=1&ToSeries=50&DAT=RNG&FD=14&FM=Oct' \
        '&FY=2014&TD=26&TM=Feb&TY=2019&VFD=Y&CSVF=TN&C=13T&Filter=N&csv' \
        '.x=28&csv.y=22'
    update_file(br_url, 'boe_hist_rates.csv')
    br = pd.read_csv('boe_hist_rates.csv')

    wb = xl.load_workbook('LIBOR.xlsx')
    la = wb["actual"]

    libor = []

    row = 1
    # loops through historic libor, subtracts base rate of period
    while la.cell(row=row, column=3).value is not None:
        ld = la.cell(row, 3).value.date()
        if(la.cell(row, 1).value == '3 Month'):
            libor_i = la.cell(row, 2).value
            if not (ld > date(2018, 3, 7) and ld < date(2018, 5, 29)):
                bank_rate = br.loc[br['DATE'] == ld.strftime(
                    '%d %b %Y')].iloc[0]['IUDBEDR']
                adj_libor = libor_i - bank_rate
                libor.append(adj_libor)
        row += 1

    return {'mean': stats.mean(libor), 'std_dev': stats.stdev(libor)}


def project_base_rates(start_date):
    # scrapes projected base rates
    br_url = 'https://public.tableau.com/views/Conditioningassumptions/CSV' \
        '.csv?:showVizHome=no'
    update_file(br_url, 'boe_proj_rates.csv')
    br = pd.read_csv('boe_proj_rates.csv')
    br = br[br['Measure'] == 'Bank rate forecast']
    br.drop('Measure', axis=1, inplace=True)
    # converts date format in scraped data to match model
    for i, r in enumerate(br['Quarter']):
        q = date(int(r[:4]), 1 + ((int(r[-1:]) - 1) * 3), 1)
        br.at[i, 'Quarter'] = q
    prev_rate = 0.0
    maturity_date = br[-1:]['Quarter'].iloc[0]
    structure = pd.DataFrame()
    payment_dates = pd.Series(pd.date_range(
        start_date, maturity_date, freq=CBD()))
    structure['dates'] = payment_dates
    structure['base rate'] = [0.0] * len(payment_dates)
    # sets the projected rate, and then sets dates out of scope as the last one
    for i, r in enumerate(structure['dates']):
        rate_set = False
        r = r.date()  # date
        prev_date = None
        for j, p in enumerate(br['Quarter']):
            if prev_date is not None:
                if (prev_date < r) and (r <= p):
                    structure.at[i, 'base rate'] = br['Value'].iloc[[j - 1]]
                    rate_set = True
            prev_date = p
    last_date = prev_date
    structure = structure.set_index('dates')
    return structure[structure.index <= pd.to_datetime(last_date)]


def historical_libor():
    prev_date = None
    wb = xl.load_workbook('LIBOR.xlsx')
    # contains historic libor data
    la = wb["actual"]
    # populates historic libor rates into structure
    # loops through all the data to find a matching date
    row = 1
    libor = {}
    while la.cell(row=row, column=3).value is not None:
        if(la.cell(row, 1).value == '3 Month'):
            libor[la.cell(row, 3).value.date()] = la.cell(row, 2).value
        row += 1
    # libor_df = pd.DataFrame.from_dict(libor)
    libor_df = pd.DataFrame(list(libor.items()), columns=['Date', 'LIBOR'])
    return libor_df


def libor_projection(libor_df, avg, std_dev):
    # scrapes projected boe base rates from boe website
    start_date = libor_df[-1:]['Date'].iloc[0]
    global BASE_RATES
    BASE_RATES = project_base_rates(start_date)
    # generates a normal distribution based on the number of future quarters
    libor_prem = np.random.normal(avg, std_dev, len(BASE_RATES)).round(4)
    libor = {}
    for i, row in enumerate(BASE_RATES.itertuples(index=True)):
        curr_date = row[0]
        base_rate = row[1]
        libor[curr_date] = base_rate + libor_prem[i]
    libor_rates = pd.DataFrame(list(libor.items()), columns=['dates', 'LIBOR'])
    BASE_RATES.reset_index(inplace=True)
    return libor_rates


def libor(num_simulations):
    libor_set = []

    avg = calculate_libor_dist()['mean']
    std_dev = calculate_libor_dist()['std_dev']
    hl = historical_libor()

    for i in range(num_simulations):
        libor_set.append(libor_projection(hl, avg, std_dev))

    return libor_set
